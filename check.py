import openpyxl
import requests
import tkinter as tk
from tkinter import filedialog
from prettytable import PrettyTable


def browse_files():
    file_paths = filedialog.askopenfilenames(
        filetypes=[('Excel Files', '*.xlsx')])
    if file_paths:
        entry_filepath.delete(0, tk.END)
        entry_filepath.insert(0, "\n".join(file_paths))


def read_excel_to_2d_list(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    data = []

    for row in worksheet.iter_rows():  # type: ignore
        field1 = row[0].value
        field2 = row[1].value
        field3 = row[2].value
        field4 = row[3].value
        data.append([field1, field2, field3, field4])

    return data


def save_2d_list_to_excel(data, file_path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for row in data:
        worksheet.append(row)  # type: ignore

    workbook.save(file_path)


excel_file_path = './usrdata.xlsx'
list_usr = []
res_usr = []


def check():
    data = read_excel_to_2d_list(excel_file_path)

    # Create a table instance
    table = PrettyTable()
    table.field_names = ['id', 'name', 'gt', 'tag']

    # Add data to the table
    for row in data:
        table.add_row(row)

    # Display the table
    print(table)

    # If you want to access the data as a 2D list directly, use 'data' variable here.
    # print('2D List:', data)
    global list_usr
    list_usr = data

    for element in list_usr:
        url = f'https://openapi.zalo.me/v2.0/oa/getprofile?data={{"user_id":\'{element[0]}\'}}'
        payload = {}
        headers = {
            'access_token': access_token
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        # print(response.text)
        print(element[1])
        if not response.json().get('data'):
            res_usr.append([element[0], element[1], element[2], 'Không có'])
            element[3] = 'Không có'
        else:
            res_usr.append([element[0], element[1], element[2], element[3]])

# check()
# save_2d_list_to_excel(res_usr, './res2.xlsx')


def browse_file():
    file_path = filedialog.askopenfilename(
        filetypes=[('Excel Files', '*.xlsx')])
    if file_path:
        entry_filepath.delete(0, tk.END)
        entry_filepath.insert(0, file_path)


def update_token():
    new_token = entry_token.get()
    if new_token:
        global access_token
        access_token = new_token
        status_label.config(text="Token updated.")


def check_single_file(file_path):
    data = read_excel_to_2d_list(file_path)

    # Create a table instance
    table = PrettyTable()
    table.field_names = ['id', 'name', 'gt', 'tag']

    # Add data to the table
    for row in data:
        table.add_row(row)

    # Display the table
    print(table)

    global list_usr
    list_usr = data

    for element in list_usr:
        url = f'https://openapi.zalo.me/v2.0/oa/getprofile?data={{"user_id":\'{element[0]}\'}}'
        payload = {}
        headers = {
            'access_token': access_token
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        print(element[1])
        if not response.json().get('data'):
            res_usr.append([element[0], element[1], element[2], 'Không có'])
            element[3] = 'Không có'
        else:
            res_usr.append([element[0], element[1], element[2], element[3]])

    # Save the results to a separate Excel file
    result_save_path = file_path.replace('.xlsx', '_result.xlsx')
    save_2d_list_to_excel(res_usr, result_save_path)

# ...


def run_check():
    file_paths = entry_filepath.get().split('\n')
    if not file_paths:
        status_label.config(text="Please select at least one Excel file.")
        return

    for file_path in file_paths:
        check_single_file(file_path)

    status_label.config(text="Check completed for selected files.")


def save_result():
    if not res_usr:
        status_label.config(text="No data to save.")
        return

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[('Excel Files', '*.xlsx')])
    if save_path:
        save_2d_list_to_excel(res_usr, save_path)
        status_label.config(text="Result saved to file.")


# Create the main application window
app = tk.Tk()
app.title("User Profile Checker")

# UI elements
label_filepath = tk.Label(app, text="Select Excel File:")
label_filepath.grid(row=0, column=0, padx=5, pady=5)

entry_filepath = tk.Entry(app, width=40)
entry_filepath.grid(row=0, column=1, padx=5, pady=5)

button_browse = tk.Button(app, text="Browse", command=browse_files)
button_browse.grid(row=0, column=2, padx=5, pady=5)

label_token = tk.Label(app, text="Enter Access Token:")
label_token.grid(row=1, column=0, padx=5, pady=5)

entry_token = tk.Entry(app, width=40)
entry_token.grid(row=1, column=1, padx=5, pady=5)

button_update_token = tk.Button(app, text="Update Token", command=update_token)
button_update_token.grid(row=1, column=2, padx=5, pady=5)

button_run = tk.Button(app, text="Run Check", command=run_check)
button_run.grid(row=2, column=0, columnspan=3, padx=5, pady=5)

button_save = tk.Button(app, text="Save Result", command=save_result)
button_save.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

status_label = tk.Label(app, text="")
status_label.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

# Set default access token
access_token = ''
entry_token.insert(0, access_token)

app.mainloop()
